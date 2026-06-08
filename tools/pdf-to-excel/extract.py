# -*- coding: utf-8 -*-
"""
절단도면 PDF → 엑셀 변환 도구 (Phase B-5)

ERP 의 [엑셀] 업로드 기능과 호환되는 형식의 엑셀을 생성합니다.

사용법:
    python extract.py <pdf_path>
    또는 PDF 파일을 run.bat 위로 드래그앤드롭

출력:
    <pdf_name>_output.xlsx (PDF 와 같은 폴더)

엑셀 컬럼:
    페이지 | 도면번호 | 부재중량(Kg) | 마킹길이(M) | 절단길이(M)

지원 양식 (ERP 시드 프리셋 v5 와 동일):
    1. 한국조선기술 NESTING (텍스트 PDF) — 사용중량(Kg) / Cut-Len / Mark-Len
    2. NC 가공도 TOTAL PART WEIGHT — TOTAL PART WEIGHT / CUTTING LEN / MARKING LEN
    3. NC 가공도 PART WEIGHT (TOTAL 없음) — PART WEIGHT / CUTTING LEN / MARKING LEN
"""

import sys
import os
import re

# Windows 콘솔 한글 출력 안전 (errors=replace — 깨져도 crash 방지)
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

try:
    import fitz  # PyMuPDF
except ImportError:
    print("[ERROR] PyMuPDF 가 설치되지 않았습니다. install.bat 을 먼저 실행하세요.")
    sys.exit(1)

# OCR 엔진 선택 — 우선순위: paddleocr → rapidocr_onnxruntime → pytesseract
# Python 3.11 venv + PaddleOCR 가 기본 구성 (install.bat 참고)
OCR_BACKEND = None
try:
    from paddleocr import PaddleOCR
    OCR_BACKEND = "paddleocr"
except ImportError:
    try:
        from rapidocr_onnxruntime import RapidOCR
        OCR_BACKEND = "rapidocr"
    except ImportError:
        try:
            import pytesseract
            OCR_BACKEND = "pytesseract"
        except ImportError:
            print("[ERROR] OCR 라이브러리가 설치되지 않았습니다. install.bat 을 먼저 실행하세요.")
            sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("[ERROR] openpyxl 이 설치되지 않았습니다. install.bat 을 먼저 실행하세요.")
    sys.exit(1)


# ────────────────────────────────────────────────────────────────
# 프리셋 (ERP 시드 v5 와 동일)
# ────────────────────────────────────────────────────────────────
PRESETS = [
    {
        "name": "한국조선기술 NESTING (텍스트 PDF)",
        "detect_keywords":  ["사용중량(Kg)", "사용중량(KG)"],
        "negative_keywords": [],
        "fields": {
            "drawing_no":  {"label": "DWG NO",       "value_pattern": r"[A-Z0-9]+NC[A-Z]\d+", "tail": 5,  "search_range": 50},
            "part_weight": {"label": "사용중량(Kg)",  "value_pattern": r"([0-9]+(?:\.[0-9]+)?)(?:\s*[Kk][Gg])?",  "search_range": 40},
            "marking_len": {"label": "Mark-Len(M)",  "value_pattern": r"([0-9]+(?:\.[0-9]+)?)(?:\s*M)?\b",        "search_range": 40},
            "cutting_len": {"label": "Cut-Len(M)",   "value_pattern": r"([0-9]+(?:\.[0-9]+)?)(?:\s*M)?\b",        "search_range": 40},
        },
    },
    {
        "name": "NC 가공도 (TOTAL PART WEIGHT)",
        "detect_keywords":  ["TOTAL PART WEIGHT"],
        "negative_keywords": [],
        "fields": {
            "drawing_no":  {"label": "DRAWING NO",        "value_pattern": r"[A-Z]+\d+(?:[-\s][A-Z0-9]+)*", "tail": 5,  "search_range": 50},
            "part_weight": {"label": "TOTAL PART WEIGHT", "value_pattern": r"([0-9]+(?:\.[0-9]+)?)(?:\s*[Kk][Gg])?",  "search_range": 40},
            "marking_len": {"label": "MARKING LEN",       "value_pattern": r"([0-9]+(?:\.[0-9]+)?)(?:\s*M)?\b",        "search_range": 40},
            "cutting_len": {"label": "CUTTING LEN",       "value_pattern": r"([0-9]+(?:\.[0-9]+)?)(?:\s*M)?\b",        "search_range": 40},
        },
    },
    {
        "name": "NC 가공도 (PART WEIGHT)",
        "detect_keywords":  ["PART WEIGHT"],
        "negative_keywords": ["TOTAL PART WEIGHT"],
        "fields": {
            "drawing_no":  {"label": "DRAWING NO",  "value_pattern": r"[A-Z0-9]+", "tail": 6, "search_range": 50},
            "part_weight": {"label": "PART WEIGHT", "value_pattern": r"([0-9]+(?:\.[0-9]+)?)(?:\s*[Kk][Gg])?",  "search_range": 40},
            "marking_len": {"label": "MARKING LEN", "value_pattern": r"([0-9]+(?:\.[0-9]+)?)(?:\s*M)?\b",        "search_range": 40},
            "cutting_len": {"label": "CUTTING LEN", "value_pattern": r"([0-9]+(?:\.[0-9]+)?)(?:\s*M)?\b",        "search_range": 40},
        },
    },
]


# ────────────────────────────────────────────────────────────────
# 라벨 매칭 알고리즘 (ERP 의 cutting-pdf-extract.ts 와 동일 로직)
# ────────────────────────────────────────────────────────────────

def normalize(s: str) -> str:
    """공백 정규화 + 대문자 + OCR 자주 혼동되는 글자 통일 (라벨 매칭 전용)"""
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s).strip().upper()
    s = s.replace("0", "O").replace("1", "I")
    return s


def detect_preset(full_text: str):
    """페이지 전체 텍스트에서 가장 많이 매칭되는 프리셋 반환"""
    upper = full_text.upper()
    best = None
    best_score = 0
    for p in PRESETS:
        neg = [k.upper() for k in p["negative_keywords"]]
        if any(kw in upper for kw in neg):
            continue
        score = sum(1 for kw in p["detect_keywords"] if kw.upper() in upper)
        if score > best_score:
            best_score = score
            best = p
    return best if best_score > 0 else None


def extract_field(full_text: str, rule: dict):
    """라벨 직후 N자 안에서 정규식 매치 + transform 적용"""
    norm_text = normalize(full_text)
    norm_label = normalize(rule["label"])
    idx = norm_text.find(norm_label)
    if idx < 0:
        return None

    rng = rule.get("search_range", 100)
    # normalize 인덱스 == 원본 인덱스 (length 동일)
    slice_text = full_text[idx + len(rule["label"]):idx + len(rule["label"]) + rng]
    m = re.search(rule["value_pattern"], slice_text)
    if not m:
        return None
    val = m.group(1) if m.groups() else m.group(0)
    if "tail" in rule:
        val = val[-rule["tail"]:]
    return val


def to_num(s):
    if s is None:
        return None
    cleaned = re.sub(r"[^\d.\-]", "", s)
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def extract_page(text: str, preset: dict):
    fields = preset["fields"]
    return {
        "drawing_no":  extract_field(text, fields["drawing_no"]),
        "part_weight": to_num(extract_field(text, fields["part_weight"])),
        "marking_len": to_num(extract_field(text, fields["marking_len"])),
        "cutting_len": to_num(extract_field(text, fields["cutting_len"])),
    }


# ────────────────────────────────────────────────────────────────
# PDF → OCR → 결과
# ────────────────────────────────────────────────────────────────

def render_page_to_image(page, scale: float = 2.5) -> bytes:
    """PDF 페이지를 PNG bytes 로 렌더. fitz 가 page.rotation 자동 적용 → 항상 정방향."""
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
    return pix.tobytes("png")


def ocr_image_bytes(ocr, img_bytes: bytes) -> str:
    """이미지 → 전체 텍스트. OCR_BACKEND 에 따라 paddleocr / rapidocr / pytesseract 사용."""
    import io
    from PIL import Image

    img = Image.open(io.BytesIO(img_bytes)).convert("RGB")

    if OCR_BACKEND == "paddleocr":
        # PaddleOCR.ocr() 결과 형식:
        #   [[ [bbox, (text, conf)], [bbox, (text, conf)], ... ]]  ← 페이지별 wrapping
        # cls=True 면 각도 분류기 사용 (회전된 텍스트도 인식)
        import numpy as np
        arr = np.array(img)
        # PaddleOCR 신/구 버전 호환: 신버전은 predict(), 구버전은 ocr()
        try:
            result = ocr.ocr(arr, cls=True)
        except TypeError:
            # cls 인자가 없는 신버전 (PaddleOCR 3.x)
            result = ocr.ocr(arr)
        if not result or not result[0]:
            return ""
        parts = []
        for line in result[0]:
            # line = [bbox, (text, confidence)]
            if len(line) >= 2 and isinstance(line[1], (tuple, list)) and len(line[1]) >= 1:
                txt = line[1][0]
                if isinstance(txt, str):
                    parts.append(txt)
        return " ".join(parts)

    elif OCR_BACKEND == "rapidocr":
        import numpy as np
        arr = np.array(img)
        result, _ = ocr(arr)
        if not result:
            return ""
        parts = []
        for item in result:
            # item = [bbox, text, confidence]
            if len(item) >= 2 and isinstance(item[1], str):
                parts.append(item[1])
        return " ".join(parts)

    elif OCR_BACKEND == "pytesseract":
        # 영문 모드 (양식 2/3 키워드 모두 영문). 한글 데이터 있으면 'kor+eng' 가능
        return pytesseract.image_to_string(img, lang="eng")

    return ""


def get_page_text(page) -> str:
    """텍스트 PDF 면 즉시 추출 (양식 1), 아니면 빈 문자열 반환 → OCR 필요"""
    txt = page.get_text("text").strip()
    return txt if len(txt) > 50 else ""


# ────────────────────────────────────────────────────────────────
# 엑셀 저장
# ────────────────────────────────────────────────────────────────

def save_excel(results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "절단도면 추출"

    headers = ["페이지", "도면번호", "부재중량(Kg)", "마킹길이(M)", "절단길이(M)"]
    ws.append(headers)

    # 헤더 스타일
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2563EB")
    center   = Alignment(horizontal="center", vertical="center")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    for r in results:
        ws.append([
            r["page"],
            r["drawing_no"] or "",
            r["part_weight"] if r["part_weight"] is not None else "",
            r["marking_len"] if r["marking_len"] is not None else "",
            r["cutting_len"] if r["cutting_len"] is not None else "",
        ])

    # 컬럼 폭
    widths = [10, 22, 16, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 메타 시트
    meta = wb.create_sheet("메타")
    meta.append(["항목", "값"])
    meta.append(["변환 행수", len(results)])
    meta.append(["변환 도구", "PaddleOCR / RapidOCR / pytesseract + PyMuPDF (Phase B-5)"])
    meta.append(["호환", "ERP cnc-erp 의 [엑셀] 업로드 버튼"])

    wb.save(output_path)


# ────────────────────────────────────────────────────────────────
# 메인
# ────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python extract.py <pdf_path>")
        sys.exit(1)

    pdf_path = sys.argv[1].strip().strip('"')
    if not os.path.exists(pdf_path):
        print(f"[ERROR] 파일이 없습니다: {pdf_path}")
        sys.exit(1)

    if not pdf_path.lower().endswith(".pdf"):
        print(f"[ERROR] PDF 파일만 지원: {pdf_path}")
        sys.exit(1)

    print("=" * 70)
    print(f"절단도면 PDF → 엑셀 변환")
    print("=" * 70)
    print(f"입력: {pdf_path}")

    # 1) PDF 열기
    print("\n[1/3] PDF 로드 중...")
    doc = fitz.open(pdf_path)
    n_pages = len(doc)
    print(f"  총 {n_pages} 페이지")

    # 2) OCR 초기화 (텍스트 PDF 만 있으면 OCR 불필요)
    ocr = None
    needs_ocr = False
    page_texts = []
    print("\n[2/3] 페이지 분석 중...")
    for i in range(n_pages):
        txt = get_page_text(doc[i])
        page_texts.append(txt)
        if not txt:
            needs_ocr = True

    if needs_ocr:
        print("  → 일부 페이지는 OCR 필요 (path-outlined PDF)")
        if OCR_BACKEND == "paddleocr":
            print("  → PaddleOCR 초기화 중... 처음 1회만 모델 다운로드 (~10MB)")
            print("  → (PaddleOCR 2.x 권장. 3.x 면 API 비호환 — README 트러블슈팅 참조)")
            # use_angle_cls=True → 회전 텍스트 인식
            # lang='en' → 영문 모드 (도면번호/라벨 모두 영문)
            # show_log=False → 진행 로그 억제
            try:
                ocr = PaddleOCR(use_angle_cls=True, lang='en', show_log=False)
            except (TypeError, ValueError):
                try:
                    # PaddleOCR 3.x 는 show_log 인자 ValueError 로 거부
                    ocr = PaddleOCR(use_angle_cls=True, lang='en')
                except (TypeError, ValueError):
                    # PaddleOCR 3.x 의 새 인자 (use_textline_orientation)
                    # 단 .ocr() 메서드 시그니처도 변경되어 호환성 제한적 — README 참조
                    ocr = PaddleOCR(use_textline_orientation=True, lang='en')
        elif OCR_BACKEND == "rapidocr":
            print("  → RapidOCR (ONNX runtime) 초기화 중... 처음 1회만 모델 다운로드")
            ocr = RapidOCR()
        elif OCR_BACKEND == "pytesseract":
            print("  → pytesseract 모드 (Tesseract 바이너리 필요)")
            ocr = None  # pytesseract 는 함수 호출, 인스턴스 X
        else:
            print(f"  [ERROR] 알 수 없는 OCR backend: {OCR_BACKEND}")
            sys.exit(1)
        print(f"  → OCR 준비 완료 (backend: {OCR_BACKEND})")
    else:
        print("  → 모든 페이지가 텍스트 PDF (OCR 불필요, 빠름)")

    # 3) 페이지마다 텍스트 → 필드 추출
    print(f"\n[3/3] 페이지별 추출 시작 ({n_pages} 페이지)")
    results = []
    skipped_no_match = 0
    skipped_empty = 0

    for i in range(n_pages):
        page_num = i + 1
        text = page_texts[i]

        if not text and ocr is not None:
            # OCR 실행
            try:
                img_bytes = render_page_to_image(doc[i], scale=2.5)
                text = ocr_image_bytes(ocr, img_bytes)
            except Exception as e:
                print(f"  [{page_num:>3}/{n_pages}] OCR 오류: {e}")
                continue

        if not text:
            print(f"  [{page_num:>3}/{n_pages}] 텍스트 없음 (skip)")
            continue

        preset = detect_preset(text)
        if preset is None:
            skipped_no_match += 1
            print(f"  [{page_num:>3}/{n_pages}] 양식 매칭 실패 (skip)")
            continue

        ext = extract_page(text, preset)
        if not any([ext["drawing_no"], ext["part_weight"], ext["marking_len"], ext["cutting_len"]]):
            skipped_empty += 1
            print(f"  [{page_num:>3}/{n_pages}] {preset['name']} → 4 필드 모두 빈 값 (skip)")
            continue

        results.append({
            "page":        page_num,
            "drawing_no":  ext["drawing_no"],
            "part_weight": ext["part_weight"],
            "marking_len": ext["marking_len"],
            "cutting_len": ext["cutting_len"],
            "preset":      preset["name"],
        })
        print(f"  [{page_num:>3}/{n_pages}] {ext['drawing_no'] or '-':<8} W={ext['part_weight'] or '-':<8} M={ext['marking_len'] or '-':<6} C={ext['cutting_len'] or '-':<6}")

    # 4) 엑셀 저장
    output_path = os.path.splitext(pdf_path)[0] + "_output.xlsx"
    save_excel(results, output_path)

    print("\n" + "=" * 70)
    print("완료")
    print("=" * 70)
    print(f"추출 행수: {len(results)} / {n_pages}")
    if skipped_no_match: print(f"양식 매칭 실패: {skipped_no_match}")
    if skipped_empty:    print(f"4 필드 빈 값:    {skipped_empty}")
    print(f"\n엑셀 저장됨: {output_path}")
    print(f"\n다음 단계:")
    print(f"  1) 엑셀 파일을 열어 추출 결과를 검토/수정하세요")
    print(f"  2) ERP 의 절단도면 PDF 탭에서 해당 PDF 의 [엑셀] 버튼을 클릭")
    print(f"  3) 위 엑셀 파일을 업로드")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[중단됨]")
    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
